"""Export normalized records and screening files."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from .models import SCHEMA_COLUMNS


def enforce_schema(df: pd.DataFrame, columns: list[str] | None = None) -> pd.DataFrame:
    cols = columns or SCHEMA_COLUMNS
    out = df.copy()
    for col in cols:
        if col not in out.columns:
            out[col] = pd.NA
    return out[cols]


def write_table(df: pd.DataFrame, path_base: Path, *, xlsx: bool = False) -> list[str]:
    path_base.parent.mkdir(parents=True, exist_ok=True)
    files = []
    csv_path = path_base.with_suffix(".csv")
    df.to_csv(csv_path, index=False, encoding="utf-8")
    files.append(str(csv_path))
    parquet_path = path_base.with_suffix(".parquet")
    df.to_parquet(parquet_path, index=False)
    files.append(str(parquet_path))
    if xlsx:
        xlsx_path = path_base.with_suffix(".xlsx")
        df.to_excel(xlsx_path, index=False)
        files.append(str(xlsx_path))
    return files


def create_screening_template(df: pd.DataFrame, output_path: Path, reasons_cfg: dict[str, Any]) -> str:
    cols = [
        "title",
        "authors",
        "year",
        "doi_normalized",
        "database_scope",
        "abstract",
        "topic_class",
        "relevance_score",
        "landing_page_url",
        "abstract_url",
        "pdf_url",
        "title_abstract_decision",
        "full_text_decision",
        "exclusion_reason",
        "reviewer",
        "review_date",
        "notes",
    ]
    out = df.copy()
    for c in cols:
        if c not in out:
            out[c] = pd.NA
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out[cols].to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    decisions = ",".join(reasons_cfg["screening_decisions"])
    reasons = ",".join(reasons_cfg["exclusion_reasons"])
    for col_name, values in [("title_abstract_decision", decisions), ("full_text_decision", decisions), ("exclusion_reason", reasons)]:
        col_idx = cols.index(col_name) + 1
        dv = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=max(len(out) + 1, 1000), column=col_idx).coordinate}")
    wb.save(output_path)
    return str(output_path)
